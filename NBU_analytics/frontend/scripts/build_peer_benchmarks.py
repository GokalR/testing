"""
Builds peer_benchmarks.json from the 4 sample SME Form 1 + Form 2 files.
Run from NBU_analytics/ root:  python frontend/scripts/build_peer_benchmarks.py
"""
import json, os, sys, io, re
from statistics import median
from openpyxl import load_workbook

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

SRC = 'SME-companies'
OUT = 'frontend/src/data/regionalStrategist/peer_benchmarks.json'

COMPANIES = [
    {'id': 'amirshoh',   'name': 'AMIRSHOH JANNATXON',          'sector': 'wholesale',
     'form1': 'AMIRSHOH JANNATXON MCHJ Баланс 2025 йиллик.xlsx',
     'form2': 'AMIRSHOH_JANNATXON_MCHJ_Молиявий_натижа_2025_йиллик.xlsx'},
    {'id': 'erkin',      'name': 'ERKIN PARVOZ NURI OLTIN',     'sector': 'services',
     'form1': 'ERKIN PARVOZ NURI OLTIN Баланс 2025 й 3 чорак.xlsx',
     'form2': 'ERKIN_PARVOZ_NURI_OLTIN_Молиявий_натижа_2025_й_3_чорак.xlsx'},
    {'id': 'osiyo',      'name': 'OSIYO NAV NIXOL FAYZ',        'sector': 'production',
     'form1': 'OSIYO NAV NIXOL FAYZ MCHJ Баланс 2025 йиллик.xltx',
     'form2': 'OSIYO_NAV_NIXOL_FAYZ_MCHJ_Молиявий_натижа_2025_йиллик.xltx'},
    {'id': 'turkiston',  'name': 'TURKISTON OBOD KELAJAGI',     'sector': 'services',
     'form1': 'TURKISTON OBOD KELAJAGI OILAVIY KORXONA форма 1  2025.xlsx',
     'form2': 'TURKISTON OBOD KELAJAGI OILAVIY KORXONA форма 2 2025.xlsx'},
]

CODE_RE = re.compile(r'^\d{3}$')

def as_num(v):
    if v is None: return None
    if isinstance(v, bool): return None
    if isinstance(v, (int, float)): return float(v)
    s = str(v).strip()
    if s in ('', 'x', 'X', '-', '—'): return None
    try:
        return float(s.replace(' ', '').replace(',', '.'))
    except ValueError:
        return None

def find_code_and_values(row, expected_value_count=2):
    """Locate the code cell and return (code, [numeric cells to the right])."""
    code_idx = None
    for j, v in enumerate(row):
        if v is None: continue
        s = str(v).strip()
        if CODE_RE.match(s):
            code_idx = j
            break
    if code_idx is None:
        return None, []
    values = []
    for v in row[code_idx+1:]:
        values.append(as_num(v))
    return str(row[code_idx]).strip(), values

def parse_form1(path):
    """Balance sheet: returns dict code -> end-period value (closing balance)."""
    wb = load_workbook(path, data_only=True)
    ws = wb['list02']
    out = {}
    for row in ws.iter_rows(values_only=True):
        code, vals = find_code_and_values(row)
        if not code: continue
        # Form 1 has 2 value columns: period start, period end.
        end = None
        nums = [v for v in vals if v is not None]
        if len(nums) >= 2:
            end = nums[1]
        elif len(nums) == 1:
            end = nums[0]
        if end is not None:
            out[code] = end
    return out

def parse_form2(path):
    """P&L: returns dict code -> current-period value (income OR expense, whichever is populated)."""
    wb = load_workbook(path, data_only=True)
    ws = wb['list02']
    out = {}
    for row in ws.iter_rows(values_only=True):
        code, vals = find_code_and_values(row)
        if not code: continue
        # Form 2 value layout after code: prior_income, prior_expense, current_income, current_expense
        # Take current period (last 2 numeric slots if available).
        nums = vals[:4] if len(vals) >= 4 else vals
        # Pad to 4
        while len(nums) < 4:
            nums.append(None)
        cur_income, cur_expense = nums[2], nums[3]
        # Pick whichever is populated (non-None). If both, sum with expense as negative sense handled by caller.
        if cur_income is not None and cur_expense is not None:
            # Some summary rows list both 0 — keep income convention.
            out[code] = cur_income
        elif cur_income is not None:
            out[code] = cur_income
        elif cur_expense is not None:
            out[code] = cur_expense
    return out

def safe_div(a, b):
    if a is None or b is None: return None
    if b == 0: return None
    return a / b

def ratios_for(f1, f2):
    g = lambda d, k: d.get(k) or 0.0

    revenue        = g(f2, '010')
    cogs           = g(f2, '020')
    gross_profit   = g(f2, '030')
    period_exp     = g(f2, '040')
    sell_exp       = g(f2, '050')
    admin_exp      = g(f2, '060')
    op_profit      = g(f2, '100')
    interest_exp   = g(f2, '180')
    pretax_profit  = g(f2, '240')
    net_profit     = f2.get('270') or f2.get('260') or f2.get('250') or pretax_profit

    fixed_assets   = g(f1, '012')
    inventory      = g(f1, '140')
    receivables    = g(f1, '210')
    cash           = g(f1, '320')
    current_assets = g(f1, '390')
    total_assets   = g(f1, '400')
    equity         = g(f1, '480')
    lt_liab        = g(f1, '490')
    lt_bank_loans  = g(f1, '570')
    st_liab        = g(f1, '600')
    st_bank_loans  = g(f1, '730')
    total_liab     = g(f1, '770')
    total_debt     = lt_bank_loans + (f1.get('580') or 0) + st_bank_loans + (f1.get('740') or 0)

    return {
        'revenue': revenue,
        'grossProfit': gross_profit,
        'operatingProfit': op_profit,
        'netProfit': net_profit,
        'totalAssets': total_assets,
        'equity': equity,
        'totalLiabilities': total_liab,
        'totalDebt': total_debt,
        'currentAssets': current_assets,
        'currentLiabilities': st_liab,
        'inventory': inventory,
        'receivables': receivables,
        'cash': cash,
        'ratios': {
            'grossMargin':     safe_div(gross_profit, revenue),
            'operatingMargin': safe_div(op_profit, revenue),
            'netMargin':       safe_div(net_profit, revenue),
            'roa':             safe_div(net_profit, total_assets),
            'roe':             safe_div(net_profit, equity),
            'currentRatio':    safe_div(current_assets, st_liab),
            'quickRatio':      safe_div(current_assets - inventory, st_liab),
            'debtToEquity':    safe_div(total_debt, equity),
            'debtToAssets':    safe_div(total_debt, total_assets),
            'assetTurnover':   safe_div(revenue, total_assets),
            'inventoryTurn':   safe_div(cogs, inventory),
            'interestCoverage':safe_div(op_profit, interest_exp),
        },
    }

def quartile(xs, q):
    xs = sorted(x for x in xs if x is not None)
    if not xs: return None
    n = len(xs)
    pos = (n - 1) * q
    lo = int(pos)
    hi = min(lo + 1, n - 1)
    frac = pos - lo
    return xs[lo] * (1 - frac) + xs[hi] * frac

def main():
    companies_out = []
    for c in COMPANIES:
        f1 = parse_form1(os.path.join(SRC, c['form1']))
        f2 = parse_form2(os.path.join(SRC, c['form2']))
        r = ratios_for(f1, f2)
        companies_out.append({'id': c['id'], 'name': c['name'], 'sector': c['sector'], **r})

    # Aggregate ratios
    ratio_keys = list(companies_out[0]['ratios'].keys())
    benchmarks = {}
    for k in ratio_keys:
        vals = [c['ratios'][k] for c in companies_out if c['ratios'][k] is not None]
        benchmarks[k] = {
            'q1':     quartile(vals, 0.25),
            'median': quartile(vals, 0.50),
            'q3':     quartile(vals, 0.75),
            'n':      len(vals),
        }

    payload = {
        'source': 'SME-companies sample (n=4, Fergana region, 2025)',
        'currency': 'UZS thousand',
        'note': 'Peer medians/quartiles from 4 Uzbek-form accounting reports. Ratios are dimensionless; absolutes in thousand UZS.',
        'benchmarks': benchmarks,
        'companies': companies_out,
    }

    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    with open(OUT, 'w', encoding='utf-8') as fh:
        json.dump(payload, fh, ensure_ascii=False, indent=2)
    print(f'Wrote {OUT}')
    print(f'Benchmarks (median):')
    for k, v in benchmarks.items():
        m = v['median']
        print(f"  {k:<18} median={m}")

if __name__ == '__main__':
    main()
