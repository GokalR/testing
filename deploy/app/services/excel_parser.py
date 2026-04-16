"""Parses Uzbek standard accounting forms (Form 1 Balance, Form 2 PnL)."""
from __future__ import annotations
import re
from io import BytesIO
from typing import Literal

from openpyxl import load_workbook

CODE_RE = re.compile(r"^\d{3}$")
FormKind = Literal["balance", "pnl"]


def _as_num(v) -> float | None:
    if v is None or isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if s in ("", "x", "X", "-", "—"):
        return None
    try:
        return float(s.replace(" ", "").replace(",", "."))
    except ValueError:
        return None


def _row_code_and_values(row):
    code_idx = None
    for j, v in enumerate(row):
        if v is None:
            continue
        if CODE_RE.match(str(v).strip()):
            code_idx = j
            break
    if code_idx is None:
        return None, []
    return str(row[code_idx]).strip(), [_as_num(v) for v in row[code_idx + 1:]]


def _pick_values_sheet(wb):
    if "list02" in wb.sheetnames:
        return wb["list02"]
    return wb.active


def parse_balance(blob: bytes) -> dict[str, float]:
    wb = load_workbook(BytesIO(blob), data_only=True)
    ws = _pick_values_sheet(wb)
    out: dict[str, float] = {}
    for row in ws.iter_rows(values_only=True):
        code, vals = _row_code_and_values(row)
        if not code:
            continue
        nums = [v for v in vals if v is not None]
        if not nums:
            continue
        out[code] = nums[1] if len(nums) >= 2 else nums[0]
    return out


def parse_pnl(blob: bytes) -> dict[str, float]:
    wb = load_workbook(BytesIO(blob), data_only=True)
    ws = _pick_values_sheet(wb)
    out: dict[str, float] = {}
    for row in ws.iter_rows(values_only=True):
        code, vals = _row_code_and_values(row)
        if not code:
            continue
        padded = (vals + [None] * 4)[:4]
        cur_income, cur_expense = padded[2], padded[3]
        if cur_income is not None:
            out[code] = cur_income
        elif cur_expense is not None:
            out[code] = cur_expense
    return out


def parse(kind: FormKind, blob: bytes) -> dict[str, float]:
    if kind == "balance":
        return parse_balance(blob)
    if kind == "pnl":
        return parse_pnl(blob)
    raise ValueError(f"Unknown form kind: {kind}")
